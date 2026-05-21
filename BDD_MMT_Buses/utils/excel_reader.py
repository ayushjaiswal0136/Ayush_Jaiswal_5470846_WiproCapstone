import openpyxl
from utils.logger import LogGen

logger = LogGen.loggen()


class ExcelReader:
    @staticmethod
    def get_test_data(filepath, sheet_name="Sheet1"):
        logger.info(f"Loading test data from: {filepath}")
        try:
            workbook = openpyxl.load_workbook(filepath, data_only=True)
            sheet = workbook[sheet_name]

            data = {}
            # Loop through all columns in the first row to get headers
            for col in range(1, sheet.max_column + 1):
                header = sheet.cell(row=1, column=col).value
                value = sheet.cell(row=2, column=col).value

                if header:
                    data[header] = value

            logger.info(f"Data successfully loaded: {data}")
            return data

        except Exception as e:
            logger.error(f"Failed to read Excel file: {str(e)}")
            raise