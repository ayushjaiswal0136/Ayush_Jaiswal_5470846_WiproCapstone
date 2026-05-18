import os

from openpyxl import load_workbook


class ExcelReader:

    @staticmethod
    def read_excel(file_name, sheet_name):

        data = []

        path = os.path.join(
            os.path.dirname(
                os.path.dirname(__file__)
            ),
            "data",
            file_name
        )

        workbook = load_workbook(path)

        sheet = workbook[sheet_name]

        headers = []

        for cell in sheet[1]:

            headers.append(cell.value)

        for row in sheet.iter_rows(
            min_row=2,
            values_only=True
        ):

            # skip empty rows
            if all(value is None for value in row):
                continue

            row_data = dict(
                zip(headers, row)
            )

            data.append(row_data)

        workbook.close()

        return data